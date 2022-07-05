import logging
import openpyxl
import os
from shutil import copy
from utils import get_output_dir, get_last_used_row_col, sort_by_quantity
from utils import update_col_widths, adjust_col_widths
from constants import VBA_ALREADY_OPEN_ERROR, SHEET_NAME, HEADERS


# GLOBAL VARIABLES
BOLD_STYLE = openpyxl.styles.Font(bold=True, name='Calibri')
FILL_NONE = openpyxl.styles.PatternFill(fill_type=None)
FILL_HIGHLIGHT = openpyxl.styles.PatternFill(fill_type='solid', fgColor='FAFA73')


class HelperFileCreate():
    '''accepts export data, sku-custom label mapping dictionaries as args, creates formatted xlsx file.
    Class does not include error handling and that should be carried out outside of this class scope

    Args: export_obj:dict - sku (key) and quantity (value int) pairs

    Main method:
    - export() - takes argument of target workbook name (path) and pushes
    sorted_export_obj accepted by class to single sheet'''
    
    def __init__(self, export_obj:dict):
        self.sorted_export_obj = sort_by_quantity(export_obj)
        self.col_widths = {}

    def export(self, wb_name:str):
        '''Creates workbook, and exports self.sorted_export_obj object to single sheet, saves new workbook'''
        wb = openpyxl.Workbook()
        self.ws = wb.active
        self.ws.freeze_panes = self.ws['A2']
        self.ws.title = SHEET_NAME
        self.fill_sheet()
        wb.save(wb_name)
        wb.close()
    
    def fill_sheet(self):
        '''pushes export object to workbook; adjusts column widths'''
        self.row_cursor = 1
        self.__fill_headers()
        self.push_data()
        adjust_col_widths(self.ws, self.col_widths)

    def __fill_headers(self):
        '''inserts 3 headers in 1:1 row. Bold style and update col widths dict for all'''
        for col, header in enumerate(HEADERS, start=1):
            self.ws.cell(self.row_cursor, col).value = header
            self.ws.cell(self.row_cursor, col).font = BOLD_STYLE
            self.col_widths = update_col_widths(self.col_widths, col, header, zero_indexed=False)
        self.row_cursor += 1

    def push_data(self):
        '''unpacks self.sorted_export_obj to self.ws sheet'''
        for sku_data in self.sorted_export_obj:
            self.ws.cell(self.row_cursor, 1).number_format = '@'
            self.ws.cell(self.row_cursor, 1).value = sku_data[0]
            self.ws.cell(self.row_cursor, 2).value = sku_data[1]
            self.col_widths = update_col_widths(self.col_widths, 1, sku_data[0], zero_indexed=False)
            self.col_widths = update_col_widths(self.col_widths, 2, str(sku_data[1]), zero_indexed=False)
            self.row_cursor += 1


class HelperFileUpdate():
    '''Reads data from wb (cleans ws when reading), merges sku data with export_obj, pushes updated data to ws and saves wb 
    
    NOTE:
    Class includes error handling, but raises Exception to hit outside error handler to close db connection and alert VBA.

    Args:
    - export_obj:dict - sku (key) and quantity (value int) pairs

    Main method:
    update_workbook() - takes argument of workbook path, reads contents, cleans sheet,
    merges current contents with incoming data in export_obj and pushes updated values'''
    
    def __init__(self, export_obj:dict):
        self.export_obj = export_obj
        self.col_widths = {}

    def update_workbook(self, inventory_file:str):
        '''main cls method. Handles reading, cleaning, formatting, merging of current and incoming data, pushes updated data'''
        try:
            # Backup and set workbook, worksheet objs
            wb = openpyxl.load_workbook(inventory_file)
            self.ws = wb[SHEET_NAME]
            self.backup_wb(inventory_file)
            
            # Read contents to dict
            current_skus = self.read_map_ws_data_to_list()
            updated_skus = self.update_sku_data(current_skus, self.export_obj)    

            # Sort by quantity and push updated values back to ws
            sorted_updated_skus = sort_by_quantity(updated_skus)
            self.write_updated_to_ws(sorted_updated_skus)
            adjust_col_widths(self.ws, self.col_widths)

            logging.info(f'Writing updated values done. Saving, closing...')
            wb.save(inventory_file)
            wb.close()
        except PermissionError as e:
            logging.critical(f'Workbook {inventory_file} already open. Err: {e}')
            print(VBA_ALREADY_OPEN_ERROR)
            raise Exception('Transition from HelperFileUpdate.updateworkbook error handling to ParseOrders.export_update_inventory_helper_file error handling')
        except Exception as e:
            logging.critical(f'Errors inside HelperFileUpdate.updateworkbook Errr: {e}. Closing wb without saving')
            wb.close()
            raise Exception('Transition from HelperFileUpdate.updateworkbook error handling to ParseOrders.export_update_inventory_helper_file error handling')

    @staticmethod
    def backup_wb(inventory_file:str):
        '''Creates a backup of workbook before new edits'''
        backup_dir = get_output_dir(client_file=False)
        backup_path = os.path.join(backup_dir, 'Inventory Reduction b4lastrun.xlsx')
        copy(inventory_file, backup_path)
        logging.info(f'Backup created at: {backup_path}, before touching {inventory_file}')

    def read_map_ws_data_to_list(self) -> dict:
        ws_limits = get_last_used_row_col(self.ws)
        assert ws_limits['max_col'] == 2, 'Template of helper file changed! Maximum column used in ws != 2'
        return self.get_ws_data_reset_highlight(ws_limits)

    def get_ws_data_reset_highlight(self, ws_limits:dict) -> dict:
        '''iterates though data rows [2:ws.max_row] in self.ws and collects sku data to dict object:
        {sku1 : qty1, sku2 : qty2, ...}
        Also deletes worksheet contents (excl headers in 1:1 row)'''
        current_sku_codes = {}
        for r in range(2, ws_limits['max_row'] + 1):
            # Reset color formatting
            self.__apply_row_highlight(r, FILL_NONE)
            # Get data
            sku, quantity = self._get_ws_row_data(r)
            self._clean_ws_row_data(r, ws_limits)
            if sku not in current_sku_codes.keys():
                current_sku_codes[sku] = quantity
            else:
                current_sku_codes[sku] += quantity
        logging.info(f'Before updating workbook has {len(current_sku_codes.keys())} distinct sku codes / data rows excl. headers')
        return current_sku_codes

    def _get_ws_row_data(self, r:int):
        '''returns sku, quantity, item from columns A,B,C in self.ws (SHEET_NAME) on r (arg) row'''
        sku = self.ws.cell(r, 1).value
        try:
            quantity = int(self.ws.cell(r, 2).value)
        except ValueError as e:
            logging.warning(f'Error converting quantity to integer, data found in wb cell: {self.ws.cell(r, 2).value}. Proceeding with string value')
            quantity = self.ws.cell(r, 2).value
        return sku, quantity

    def _clean_ws_row_data(self, r:int, ws_limits:dict):
        '''deletes row r contents in self.ws worksheet (SHEET_NAME)'''
        for col in range(1, ws_limits['max_col'] + 1):
            self.ws.cell(r, col).value = None

    def update_sku_data(self, current_skus:dict, export_obj:dict) -> dict:
        '''updates current_skus dict with values from loaded data (export_obj), returns same form obj:
        {sku1 : qty1, sku2 : qty2, ...}'''
        for sku in export_obj:
            if sku not in current_skus:
                logging.debug(f'Adding a new sku code: {sku}. Q-ty: {export_obj[sku]}')
                current_skus[sku] = export_obj[sku]
            else:
                logging.debug(f'Updating quantity for code: {sku}. Previous quantity: {current_skus[sku]}, adding: {export_obj[sku]}')
                current_skus[sku] += export_obj[sku]
        return current_skus

    def write_updated_to_ws(self, sorted_updated_skus:list):
        '''write sorted_updated_skus list of tuples to rows below header'''
        for row_cursor, sku_data in enumerate(sorted_updated_skus, start=2):
            self.ws.cell(row_cursor, 1).number_format = '@'    
            self.ws.cell(row_cursor, 1).value = sku_data[0]
            self.ws.cell(row_cursor, 2).value = sku_data[1]
            self.col_widths = update_col_widths(self.col_widths, 1, sku_data[0], zero_indexed=False)
            self.col_widths = update_col_widths(self.col_widths, 2, str(sku_data[1]), zero_indexed=False)

    def __apply_row_highlight(self, r:int, highlight_style=FILL_HIGHLIGHT):
        '''applies passed formatting style on self.ws r row, hardcoded 4 columns fill'''
        for c in range(1, 4):
            self.ws.cell(r, c).fill = highlight_style


if __name__ == "__main__":
    pass