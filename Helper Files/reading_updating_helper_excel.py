import openpyxl
import os
from shutil import copy
from amzn_parser_utils import get_output_dir, get_last_used_row_col

from parse_orders import EXPORT_FILE
from helper_file import SUMMARY_SHEET_NAME
from pprint import pprint

TEST_EXPORT_OBJ = {'EX200': {
                        'item': 'GOLDEN UNIVERSAL TAROT DECK [Karten] [2013] Lo Scarabeo',
                        'quantity': 15},
                    'EX198': {
                        'item': 'As Above Deck: Book of Shadows Tarot, Volume 1 [Karten] [2013] Lo Scarabeo',
                        'quantity': 5},
                    'MySampleSKU': {
                        'item': 'This is a test item title',
                        'quantity': 9}
                    }


def backup_wb():
    output_dir = get_output_dir(client_file=False)
    target_path = os.path.join(output_dir, 'Inventory Reduction b4lastrun.xlsx')
    copy(EXPORT_FILE, target_path)
    print(f'Backup created at: {target_path}, before touching {EXPORT_FILE}')

def read_ws_to_sku_dict(ws:object):
    try:
        ws_limits = get_last_used_row_col(ws)
        assert ws_limits['max_col'] == 3, 'Template of helper file changed! Maximum column used in ws != 3'
        current_sku_codes = get_ws_data(ws, ws_limits)
        return current_sku_codes
    except Exception as e:
        print(f'Error handling workbook. Error: {e}')
        print('Saving, closing workbook...')


def get_ws_data(ws:object, ws_limits:dict) -> dict:
    '''iterates though data rows [2:ws.max_row] and collects sku data to dict object:
    {sku1:{'quantity':2, 'item':'item_title1'}, sku2:{'quantity':4, 'item':'item_title2'}, ...}'''
    current_sku_codes = {}
    try:
        for r in range(2, ws_limits['max_row'] + 1):
            sku, quantity, item = get_ws_row_data(ws, r)
            clean_ws_row_data(ws, r, ws_limits)
            if sku not in current_sku_codes.keys():
                current_sku_codes[sku] = {'item':item, 'quantity':quantity}
            else:
                current_sku_codes[sku]['quantity'] += quantity
        print('Before returning current sku codes. Sheet is empty. Headers only. Lets check that')
        new_limits = get_last_used_row_col(ws)
        print(f'Printing new ws limits dict: {new_limits}')
        return current_sku_codes
    except Exception as e:
        print(f'Error collecting data from ws. Error: {e}')
        print('Saving, closing workbook...')

def get_ws_row_data(ws:object, r:int):
    sku = ws.cell(r, 1).value
    try:
        quantity = int(ws.cell(r, 2).value)
    except ValueError as e:
        print(f'Error converting quantity to integer, data found in wb cell: {ws.cell(r, 2).value}. Proceeding with string value')
        quantity = ws.cell(r, 2).value
    item = ws.cell(r, 3).value
    return sku, quantity, item

def clean_ws_row_data(ws:object, r:int, ws_limits:dict):
    '''deletes row r contents in passed ws'''
    for col in range(1, ws_limits['max_col'] + 1):
        ws.cell(r, col).value = None

def update_sku_data(current_sku_codes:dict, loaded_sku_data:dict) -> dict:
    '''updates current_sku_codes dict with values from loaded data, returns same form obj:
    {sku1:{'quantity':2, 'item':'item_title1'}, sku2:{'quantity':4, 'item':'item_title2'}, ...}'''
    for sku in loaded_sku_data.keys():

        if sku not in current_sku_codes.keys():
            print(f'Adding a new sku code: {sku}')
            current_sku_codes[sku] = {'item':loaded_sku_data[sku]['item'], 'quantity':loaded_sku_data[sku]['quantity']}
        else:
            print(f'Updating quantity for code: {sku}.')
            print(f"Previous quantity: {current_sku_codes[sku]['quantity']}, adding: {loaded_sku_data[sku]['quantity']}")

            current_sku_codes[sku]['quantity'] += loaded_sku_data[sku]['quantity']
            print(f"Updated quantity: {current_sku_codes[sku]['quantity']}")

    return current_sku_codes

def write_updated_to_ws(ws:object, updated_sku_data:dict):
    '''write updated dict to rows below header'''
    for row_cursor, sku in enumerate(updated_sku_data.keys(), start=2):
        ws.cell(row_cursor, 1).value = sku
        ws.cell(row_cursor, 2).value = updated_sku_data[sku]['quantity']
        ws.cell(row_cursor, 3).value = updated_sku_data[sku]['item']

def run():
    # backup_wb()
    WB_FILE = 'Amazon Inventory Reduction - Copy.xlsx'
    FILE_USED = WB_FILE
    # Open WB
    wb = openpyxl.load_workbook(FILE_USED)
    ws = wb[SUMMARY_SHEET_NAME]
    
    # Read contents to object
    data_in_wb = read_ws_to_sku_dict(ws)
    # print('------------------')
    # pprint(data_in_wb)
    # print('------------------')

    updated_sku_codes_data = update_sku_data(data_in_wb, TEST_EXPORT_OBJ)
    print(f'Updated sku codes object has {len(updated_sku_codes_data.keys())} entries')
    print(f'Start writing it to ws...')
    write_updated_to_ws(ws, updated_sku_codes_data)
    print(f'Writing updated values done. Saving, closing...')
    # Close WB
    wb.save(FILE_USED)
    wb.close()

def test_copy_save():
    WB_FILE = 'Amazon Inventory Reduction - Copy.xlsx'
    wb = openpyxl.load_workbook(WB_FILE)
    ws = wb.active
    max_row, max_col = get_last_used_row_col(ws)
    print(f'Detected last used row: {max_row}. Max column: {max_col}')

    print('Deleting 14 row contents')
    for col in range(1, max_col + 1):
        ws.cell(14, col).value = None
    print('About cleared, Sir')
    # max_row = get_ws_last_used_row(ws)
    # max_row = ws.max_row
    # ws['A1'].value = 'sku'

    # ws['A1'].value = 'THIS WONT BE SAVED'
    wb.save(WB_FILE)
    wb.close()
    print('saved, closed')

def test_sorting():
    sorted_export_obj = sort_labels(TEST_EXPORT_OBJ)
    print(type(sorted_export_obj))
    print('hey, seems finished')

def sort_labels(labels:dict) -> list:
    '''sorts default dict by descending quantities. Returns list of tuples'''
    return sorted(labels.items(), key=lambda sku_dict: sku_dict[1]['quantity'], reverse=True)

if __name__ == "__main__":
    # run()
    # test_copy_save()
    test_sorting()