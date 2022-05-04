from constants import VBA_ERROR_ALERT, COUNTRY_CODES, VBA_KEYERROR_ALERT, EXPORT_FILE
from openpyxl.utils import get_column_letter
from datetime import datetime
import platform
import logging
import shutil
import json
import sys
import os
import re


def get_level_up_abspath(absdir_path):
    '''returns directory absolute path one level up from passed abs path'''
    return os.path.dirname(absdir_path)

def get_output_dir(client_file=True):
    '''returns target dir for output files depending on execution type (.exe/.py) and file type (client/systemic)'''
    # pyinstaller sets 'frozen' attr to sys module when compiling
    if getattr(sys, 'frozen', False):
        curr_folder = os.path.dirname(sys.executable)
    else:
        curr_folder = os.path.dirname(os.path.abspath(__file__))
    return get_level_up_abspath(curr_folder) if client_file else curr_folder

def file_to_binary(abs_fpath:str):
    '''returns binary data for file'''
    try:
        with open(abs_fpath, 'rb') as f:
            bfile = f.read()
        return bfile
    except FileNotFoundError as e:
        print(f'file_to_binary func got arg: {abs_fpath}; resulting in error: {e}')
        return None

def recreate_txt_file(abs_fpath:str, binary_data):
    '''outputs a file from given binary data'''
    try:
        with open(abs_fpath, 'wb') as f:
            f.write(binary_data)
    except TypeError:
        print(f'Expected binary when writing contents to file {abs_fpath}')

def is_windows_machine() -> bool:
    '''returns True if machine executing the code is Windows based'''
    machine_os = platform.system()
    return True if machine_os == 'Windows' else False

def alert_vba_date_count(filter_date, orders_count):
    '''Passing two variables for VBA to display for user in message box'''
    print(f'FILTER_DATE_USED: {filter_date}')
    print(f'SKIPPING_ORDERS_COUNT: {orders_count}')

def alert_VBA_duplicate_mapping_sku(sku_code:str):
    '''duplicate SKU code found when reading mapping xlsx, alert VBA'''
    print(f'DUPLICATE SKU IN MAPPING: {sku_code}')

def get_datetime_obj(date_str):
    '''returns tz-naive datetime obj from date string. Designed to work with str format: 2020-04-16T10:07:16+00:00'''
    try:
        return datetime.fromisoformat(date_str).replace(tzinfo=None)
    except ValueError:
        # Attempt to handle wrong/new date format here
        logging.warning(f'Change in format detected! Previous format: 2020-04-16T10:07:16+00:00. Current: {date_str} Attempting to parse string...')
        try:
            date_str_split = date_str.split('T')[0]
            return datetime.fromisoformat(date_str_split)
        except ValueError:
            logging.critical(f'Unable to create datetime from date string: {date_str}. Terminating.')
            print(VBA_ERROR_ALERT)
            sys.exit()

def simplify_date(date_str : str) -> str:
    '''returns a simplified date format: YYYY-MM-DD from rawformat 2020-04-16T06:53:44+00:00'''
    try:
        date = get_datetime_obj(date_str).date()
        return date.strftime('%Y-%m-%d')
    except ValueError:
        logging.warning(f'Unable to return simplified version of date: {date_str}. Returning raw format instead')
        return date_str

def col_to_letter(col : int, zero_indexed=True) -> str:
    '''returns column letter from worksheet column index'''
    if zero_indexed:
        col += 1
    return get_column_letter(col)

def get_last_used_row_col(ws:object):
    '''returns dictionary containing max_row and max_col as integers - last used row and column in passed openpyxl worksheet'''
    row = ws.max_row
    while row > 0:
        cells = ws[row]
        if all([cell.value is None for cell in cells]):
            row -= 1
        else:
            break
    if row == 0:
        return {'max_row' : 0, 'max_col' : 0}

    column = ws.max_column
    while column > 0:
        cells = next(ws.iter_cols(min_col=column, max_col=column, max_row=row))
        if all([cell.value is None for cell in cells]):
            column -= 1
        else:
            break
    return {'max_row' : row, 'max_col' : column}

def dump_to_json(export_obj, json_fname:str) -> str:
    '''exports export_obj to json file. Returns path to crated json'''
    output_dir = get_output_dir(client_file=False)
    json_path = os.path.join(output_dir, json_fname)
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(export_obj, f, indent=4)
    return json_path

def read_json_to_obj(json_file_path:str):
    '''reads json file and returns python object'''
    with open(json_file_path, 'r', encoding='utf-8') as f:
        orders = json.load(f)
    return orders

def sort_by_quantity(sku_qties:dict) -> list:
    '''sorts {'sku1': qty1, 'sku2': qty2, ...} dict
    by descending quantities. Returns list of tuples:
    
    [('sku1', qty_max), ('sku2', qty), ..., ('sku2', qty_min)]'''
    return sorted(sku_qties.items(), key=lambda x: x[1], reverse=True)

def get_country_code(country:str) -> str:
    '''using COUNTRY_CODES dict, returns 2 letter str for country if len(country) > 2. Called from main'''
    try:
        if len(country) > 2:
            country_code = COUNTRY_CODES[country.upper()]
            return country_code
        else:
            return country
    except KeyError as e:
        logging.critical(f'Failed to get country code for: {country}. Err:{e}. Alerting VBA, terminating immediately')
        print(VBA_ERROR_ALERT)
        sys.exit()

def split_sku(split_sku:str, sales_channel:str) -> list:
    '''splits sku string on ',' and ' + ' into list of skus for Etsy.
    example input: '1 vnt. 1040830 + 1 vnt. 1034630,1 vnt. T1147'
    return value: ['1 vnt. 1040830', '1 vnt. 1034630', '1 vnt. T1147']
    
    for Amazon, only splits multilistings on plus ' + ' string'''
    if sales_channel == 'Etsy':
        plus_comma_split = [sku_sublist.split(',') for sku_sublist in split_sku.split(' + ')]
        return [sku for sku_sublist in plus_comma_split for sku in sku_sublist]
    else:
        return split_sku.split(' + ')

def create_src_file_backup(target_file_abs_path:str, backup_fname_prefix:str) -> str:
    '''returns abspath of created file backup'''
    src_files_folder = get_src_files_folder()
    _, backup_ext = os.path.splitext(target_file_abs_path)
    backup_abspath = get_backup_f_abspath(src_files_folder, backup_fname_prefix, backup_ext)
    shutil.copy(src=target_file_abs_path, dst=backup_abspath)
    logging.info(f'Backup created at: {backup_abspath}')
    return backup_abspath

def get_src_files_folder():
    output_dir = get_output_dir(client_file=False)
    target_dir = os.path.join(output_dir, 'src files')
    if not os.path.exists(target_dir):
        os.mkdir(target_dir)
        logging.debug(f'src files directory inside Helper files has been recreated: {target_dir}')
    return target_dir

def get_backup_f_abspath(src_files_folder:str, backup_fname_prefix:str, ext:str) -> str:
    '''returns abs path for backup file. fname format: backup_fname_prefix-YY-MM-DD-HH-MM.ext'''
    timestamp = datetime.now().strftime('%y-%m-%d %H-%M')
    backup_fname = f'{backup_fname_prefix} {timestamp}{ext}'
    return os.path.join(src_files_folder, backup_fname)

def delete_file(file_abspath:str):
    '''deletes file located in file_abspath'''
    try:
        os.remove(file_abspath)
    except FileNotFoundError:
        logging.warning(f'Tried deleting file: {file_abspath}, but apparently human has taken care of it first. (File not found)')
    except Exception as e:
        logging.warning(f'Unexpected err: {e} while flushing db old records, deleting file: {file_abspath}')

def get_order_quantity(order:dict, proxy_keys:dict) -> int:
    '''returns 'quantity-purchased' order key value as integer'''
    try:
        return int(order[proxy_keys['quantity-purchased']])
    except KeyError:
        logging.critical(f'Failed to retrieve order quantity for order: {order}. Proxy keys: {proxy_keys}. Returning 1')
        print(VBA_KEYERROR_ALERT)
        return 1
    except ValueError:
        logging.critical(f'Failed to convert order quantity for order: {order}. Proxy keys: {proxy_keys}. Returning 1')
        print(VBA_ERROR_ALERT)
        return 1

def get_inner_qty_sku(original_code:str, quantity_pattern:str):
    '''returns recognized internal quantity from passed regex pattern: quantity_pattern inside original_code arg and simplified code
    two examples: from codes: '(3 vnt.) CR2016 5BL 3V VINNIC LITHIUM' / '1 vnt. 1034630' ->
    return values are: 3, 'CR2016 5BL 3V VINNIC LITHIUM' / 1, '1034630' '''
    try:
        quantity_str = re.findall(quantity_pattern, original_code)[0]
        inner_quantity = int(re.findall(r'\d+', quantity_str)[0])
        inner_code = original_code.replace(quantity_str, '')
        return inner_quantity, inner_code
    except:
        return 1, original_code

def export_invalid_order_ids(invalid_orders:list, proxy_keys:dict, invalid_orders_fpath:str):
    '''exports etsy / amazon order IDs to txt file'''
    with open(invalid_orders_fpath, 'w') as f:
        f.write(f'Order ID(s), that were not included in {EXPORT_FILE}:\n\n')
        for order in invalid_orders:
            f.write(f'{order[proxy_keys["order-id"]]}\n\n')

def update_col_widths(col_widths:dict, col:int, cell_value:str, zero_indexed=True):
    '''runs on each cell. Forms a dictionary {'A':30, 'B':15...} for max column widths in worksheet (width as length of max cell)'''
    col_letter = col_to_letter(col, zero_indexed=zero_indexed)
    if col_letter in col_widths:
        # check for length, update if current cell length exceeds current entry for column
        if len(cell_value) > col_widths[col_letter]:
            col_widths[col_letter] = len(cell_value)
    else:
        col_widths[col_letter] = len(cell_value)
    return col_widths

def adjust_col_widths(ws:object, col_widths:dict):
    '''iterates over {'A':30, 'B':40, 'C':35...} dict to resize worksheets' column widths'''
    for col_letter in col_widths:
        adjusted_width = col_widths[col_letter] + 4
        ws.column_dimensions[col_letter].width = adjusted_width


if __name__ == "__main__":
    pass